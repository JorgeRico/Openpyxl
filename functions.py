import openpyxl
from openpyxl import Workbook
import random

class Functions:

    def __init__(self):
        self.titles   = [ 'sheet 1', 'sheet 2' ]
        self.filename = "sample.xlsx"

    # save info on xlsx file
    def save(self):
        workbook = Workbook()

        for title in self.titles:
            worksheet = workbook.create_sheet(title)
            self.saveSingleCell(worksheet, 'A1', 'test value')
            self.saveDataRow(worksheet, [random.randint(1, 55), random.randint(1, 55), random.randint(1, 55), random.randint(1, 55)])
            self.saveDataRow(worksheet, [random.randint(1, 55), random.randint(1, 55), random.randint(1, 55), random.randint(1, 55)])
            self.saveDataRow(worksheet, [random.randint(1, 55), random.randint(1, 55), random.randint(1, 55), random.randint(1, 55)])

        # Save the file
        workbook.save(self.filename)

    # read info from xlsx file
    def read(self):
        book = openpyxl.load_workbook(self.filename)

        for sheetname in book.sheetnames: 
            sheet = book[sheetname]

            allrows = []
            for row in sheet.values:
                rowValues = []
                for value in row:
                    rowValues.append(value)
                allrows.append(rowValues)

            # print only not empty sheets
            if len(allrows) > 0:
                self.printInfo(sheetname, allrows)

    # print sheet info
    def printInfo(self, sheetname, rows):
        print(' - sheetname: %s : %s' %(sheetname, rows))

    # save info on a single cell
    def saveSingleCell(self, worksheet, cellId, value):
        worksheet[cellId] = value
        
    # save info in a row    
    def saveDataRow(self, worksheet, items):
        worksheet.append(items)

    # read single cell info
    def readSingleCell(self, sheet, cellId):
        value = sheet[cellId].value
        
        return value